"""
Stock Rating Engine — Data Requirements Report Generator
Outputs: rating_data_requirements.pdf and rating_data_requirements.xlsx
Run from backend/ directory: python generate_rating_data_requirements.py
"""

import os
from datetime import datetime

# ── Data Table Definition ──────────────────────────────────────────────────────
# Columns:
#   field, db_table, db_column, source, update_freq, mandatory,
#   bucket, max_pts, sector_applicability, field_type,
#   missing_default, missing_impact, bank_ignored, notes

FIELDS = [
    # ─── TREND BUCKET (max 20) ────────────────────────────────────────────────
    {
        "field": "Current Market Price (CMP)",
        "db_table": "price_cache",
        "db_column": "price",
        "source": "PriceCache (fetched via yfinance / Finvasia)",
        "update_freq": "Daily (market hours)",
        "mandatory": "YES",
        "bucket": "Trend",
        "max_pts": 20,
        "sector": "All",
        "field_type": "Raw",
        "missing_default": "Score = 4 (minimum)",
        "missing_impact": "Trend score floors at 4; Valuation/Growth also break",
        "bank_ignored": "No",
        "notes": "Core field — all derived fields depend on this",
    },
    {
        "field": "50-Day Moving Average (MA50)",
        "db_table": "price_cache",
        "db_column": "ma50",
        "source": "PriceCache (computed from historical feed)",
        "update_freq": "Daily",
        "mandatory": "Recommended",
        "bucket": "Trend",
        "max_pts": 20,
        "sector": "All",
        "field_type": "Cached/Derived",
        "missing_default": "Score = 4 (minimum)",
        "missing_impact": "Trend score cannot distinguish near-term momentum",
        "bank_ignored": "No",
        "notes": "price > ma50 > ma200 → full 20 pts",
    },
    {
        "field": "200-Day Moving Average (MA200)",
        "db_table": "price_cache",
        "db_column": "ma200",
        "source": "PriceCache (computed from historical feed)",
        "update_freq": "Daily",
        "mandatory": "Recommended",
        "bucket": "Trend",
        "max_pts": 20,
        "sector": "All",
        "field_type": "Cached/Derived",
        "missing_default": "Score = 4 (minimum)",
        "missing_impact": "Cannot confirm long-term bullish setup",
        "bank_ignored": "No",
        "notes": "Golden cross (ma50 > ma200) adds significant score",
    },

    # ─── VALUATION BUCKET (max 25) ────────────────────────────────────────────
    {
        "field": "Trailing EPS",
        "db_table": "price_cache",
        "db_column": "eps",
        "source": "PriceCache (Finvasia / yfinance)",
        "update_freq": "Quarterly (on earnings)",
        "mandatory": "YES — for P/E computation",
        "bucket": "Valuation + Growth",
        "max_pts": 25,
        "sector": "All",
        "field_type": "Raw",
        "missing_default": "pe_1y/pe_2y = None → score_a = 2 (worst)",
        "missing_impact": "Valuation score capped at 2 (Part A) + absolute PE only",
        "bank_ignored": "No",
        "notes": "Used to derive forward P/E if forward_eps absent",
    },
    {
        "field": "Trailing P/E (pe_current)",
        "db_table": "price_cache",
        "db_column": "pe",
        "source": "PriceCache (price / eps, or direct feed)",
        "update_freq": "Daily (price changes; EPS quarterly)",
        "mandatory": "YES",
        "bucket": "Valuation + Growth",
        "max_pts": 25,
        "sector": "All",
        "field_type": "Derived",
        "missing_default": "Score B = 0; Growth = 2",
        "missing_impact": "No absolute valuation scoring; Growth defaults to minimum",
        "bank_ignored": "No",
        "notes": "pe < 15 → 10 pts; 15–25 → 7 pts; 25–40 → 4 pts; >40 → 0 pts",
    },
    {
        "field": "Forward EPS (Yahoo forwardEps)",
        "db_table": "price_cache",
        "db_column": "forward_eps",
        "source": "Yahoo Finance API (earningsGrowth field)",
        "update_freq": "Weekly / On analyst revision",
        "mandatory": "Optional (preferred)",
        "bucket": "Valuation + Growth",
        "max_pts": 15,
        "sector": "All",
        "field_type": "Raw (external)",
        "missing_default": "Falls back to eps * (1 + growth_rate)",
        "missing_impact": "1Y forward P/E less accurate without analyst consensus",
        "bank_ignored": "No",
        "notes": "pe_1y = price / forward_eps if available",
    },
    {
        "field": "Yahoo Earnings Growth Rate",
        "db_table": "price_cache",
        "db_column": "earnings_growth",
        "source": "Yahoo Finance API (earningsGrowth, decimal e.g. 0.15)",
        "update_freq": "Weekly",
        "mandatory": "Optional (preferred fallback)",
        "bucket": "Valuation + Growth",
        "max_pts": 15,
        "sector": "All",
        "field_type": "Raw (external)",
        "missing_default": "Falls back to eps_growth (user-saved %)",
        "missing_impact": "pe_2y derivation uses less reliable estimate",
        "bank_ignored": "No",
        "notes": "Priority: earnings_growth > eps_growth > 10% default",
    },
    {
        "field": "User EPS Growth Rate (%)",
        "db_table": "price_cache",
        "db_column": "eps_growth",
        "source": "User-editable in Equity Analytics modal",
        "update_freq": "On demand (user sets once)",
        "mandatory": "Optional (last-resort fallback)",
        "bucket": "Valuation + Growth",
        "max_pts": 15,
        "sector": "All",
        "field_type": "Raw (user input)",
        "missing_default": "Defaults to 10% if both Yahoo and user values absent",
        "missing_impact": "Minimal if Yahoo data present",
        "bank_ignored": "No",
        "notes": "3rd priority in growth-rate chain",
    },
    {
        "field": "1Y Forward P/E (pe_1y)",
        "db_table": "Computed at runtime",
        "db_column": "—",
        "source": "Derived: price / forward_eps  OR  price / (eps × (1+g))",
        "update_freq": "On rating compute",
        "mandatory": "Derived",
        "bucket": "Valuation + Growth",
        "max_pts": 15,
        "sector": "All",
        "field_type": "Derived",
        "missing_default": "score_a = 2 (minimum)",
        "missing_impact": "Forward PE trend unavailable; Growth misses compression signal",
        "bank_ignored": "No",
        "notes": "Requires EPS > 0 and price > 0",
    },
    {
        "field": "2Y Forward P/E (pe_2y)",
        "db_table": "Computed at runtime",
        "db_column": "—",
        "source": "Derived: price / (eps × (1+g)²)",
        "update_freq": "On rating compute",
        "mandatory": "Derived",
        "bucket": "Valuation + Growth",
        "max_pts": 5,
        "sector": "All",
        "field_type": "Derived",
        "missing_default": "score_a partial; growth = 14 or 2 (no 2Y)",
        "missing_impact": "Cannot confirm strong 2Y earnings compression for full 20 pts Growth",
        "bank_ignored": "No",
        "notes": "pe_2y < pe_1y < pe_current → both Valuation and Growth get max",
    },

    # ─── PROFITABILITY BUCKET (max 35) ────────────────────────────────────────
    {
        "field": "ROE – 3-Year Average (%)",
        "db_table": "stock_profitability_summary",
        "db_column": "roe_3y_avg",
        "source": "Computed by profitability service (annual reports / Screener)",
        "update_freq": "Annual (post Q4 results)",
        "mandatory": "YES — profitability score depends on this",
        "bucket": "Profitability",
        "max_pts": 10,
        "sector": "All",
        "field_type": "Derived",
        "missing_default": "roe_3y treated as 0 → score = 4 (min tier)",
        "missing_impact": "Profitability score artificially low; stock may be underrated",
        "bank_ignored": "No — redistributed: max 19 pts for Banks",
        "notes": "Non-fin: >18% → 10; ≥14% → 7; <14% → 4. Bank: >18% → 19; ≥14% → 13; else → 8",
    },
    {
        "field": "ROCE – 3-Year Average (%)",
        "db_table": "stock_profitability_summary",
        "db_column": "roce_3y_avg",
        "source": "Computed by profitability service (annual reports / Screener)",
        "update_freq": "Annual (post Q4 results)",
        "mandatory": "YES for Non-Financial; IGNORED for Banks/NBFCs",
        "bucket": "Profitability",
        "max_pts": 9,
        "sector": "Non-Financial only",
        "field_type": "Derived",
        "missing_default": "roce_3y treated as 0 → score = 3 (min tier)",
        "missing_impact": "Up to 9 pts lost; profitability looks poor",
        "bank_ignored": "YES — ROCE not applicable to banks (asset structure differs)",
        "notes": "Non-fin: >16% → 9; ≥12% → 6; <12% → 3. Skipped entirely for banks",
    },
    {
        "field": "ROE – Current Year (%)",
        "db_table": "stock_profitability_summary",
        "db_column": "roe_3y_avg (used as proxy)",
        "source": "3Y avg used as current proxy (dedicated field absent)",
        "update_freq": "Annual",
        "mandatory": "Inferred from 3Y avg",
        "bucket": "Profitability",
        "max_pts": 8,
        "sector": "All",
        "field_type": "Derived (proxy)",
        "missing_default": "Falls back to 0 → score = 3 (min tier)",
        "missing_impact": "Up to 5 pts lost in non-bank path",
        "bank_ignored": "No — redistributed: max 16 pts for Banks",
        "notes": "TODO: add dedicated roe_current column to profitability table",
    },
    {
        "field": "ROCE – Current Year (%)",
        "db_table": "stock_profitability_summary",
        "db_column": "roce_3y_avg (used as proxy)",
        "source": "3Y avg used as current proxy (dedicated field absent)",
        "update_freq": "Annual",
        "mandatory": "Inferred from 3Y avg (Non-Financial only)",
        "bucket": "Profitability",
        "max_pts": 8,
        "sector": "Non-Financial only",
        "field_type": "Derived (proxy)",
        "missing_default": "Falls back to 0 → score = 3 (min tier)",
        "missing_impact": "Up to 5 pts lost",
        "bank_ignored": "YES",
        "notes": "TODO: add dedicated roce_current column",
    },
    {
        "field": "is_bank flag",
        "db_table": "stock_profitability_summary",
        "db_column": "is_bank",
        "source": "Set during profitability computation (sector detection)",
        "update_freq": "One-time / On sector reclassification",
        "mandatory": "YES — determines scoring path",
        "bucket": "Profitability",
        "max_pts": "N/A (flag)",
        "sector": "All",
        "field_type": "Raw (flag)",
        "missing_default": "Treated as Non-Financial (False)",
        "missing_impact": "Banks scored with ROCE = 0 → lose up to 12 pts",
        "bank_ignored": "N/A",
        "notes": "Hard-coded NBFC list: BAJAJFIN, BAJFINANCE, HDFC, LIC, MUTHOOTFIN, PFC, REC, etc.",
    },
]

# ─── Scoring Thresholds Summary ───────────────────────────────────────────────
THRESHOLDS = [
    # Trend
    ("Trend: Fully Bullish",       "price > ma50 > ma200",   20, "All"),
    ("Trend: Long-term bullish",   "price > ma200, price < ma50", 14, "All"),
    ("Trend: Short-term bullish",  "price < ma50, price > ma200", 10, "All"),
    ("Trend: Bearish",             "price < ma200",           4, "All"),
    # Valuation Part A
    ("Valuation A: PE compressing 2Y", "pe_2y < pe_1y < pe_current", 15, "All"),
    ("Valuation A: PE compressing 1Y", "pe_1y < pe_current",          10, "All"),
    ("Valuation A: PE flat",           "|pe_1y - pe_current| < 1",    6, "All"),
    ("Valuation A: PE expanding",      "pe_1y >= pe_current + 1",     2, "All"),
    # Valuation Part B
    ("Valuation B: Deep Value",   "pe_current < 15",                  10, "All"),
    ("Valuation B: Fair Value",   "15 ≤ pe_current ≤ 25",             7, "All"),
    ("Valuation B: Expensive",    "25 < pe_current ≤ 40",             4, "All"),
    ("Valuation B: Very Expensive","pe_current > 40",                  0, "All"),
    # Profitability (Non-Fin)
    ("Profitability: ROE curr >20%",  "roe_current > 20",  8, "Non-Financial"),
    ("Profitability: ROE curr ≥15%",  "roe_current ≥ 15",  6, "Non-Financial"),
    ("Profitability: ROE curr <15%",  "roe_current < 15",  3, "Non-Financial"),
    ("Profitability: ROCE curr >18%", "roce_current > 18", 8, "Non-Financial"),
    ("Profitability: ROCE curr ≥14%", "roce_current ≥ 14", 6, "Non-Financial"),
    ("Profitability: ROCE curr <14%", "roce_current < 14", 3, "Non-Financial"),
    ("Profitability: ROE 3Y >18%",   "roe_3y > 18",       10, "Non-Financial"),
    ("Profitability: ROE 3Y ≥14%",   "roe_3y ≥ 14",        7, "Non-Financial"),
    ("Profitability: ROE 3Y <14%",   "roe_3y < 14",        4, "Non-Financial"),
    ("Profitability: ROCE 3Y >16%",  "roce_3y > 16",        9, "Non-Financial"),
    ("Profitability: ROCE 3Y ≥12%",  "roce_3y ≥ 12",        6, "Non-Financial"),
    ("Profitability: ROCE 3Y <12%",  "roce_3y < 12",        3, "Non-Financial"),
    # Bank path
    ("Profitability (Bank): ROE curr >20%", "roe_current > 20", 16, "Bank/NBFC"),
    ("Profitability (Bank): ROE curr ≥15%", "roe_current ≥ 15", 12, "Bank/NBFC"),
    ("Profitability (Bank): ROE curr <15%", "roe_current < 15",  6, "Bank/NBFC"),
    ("Profitability (Bank): ROE 3Y >18%",   "roe_3y > 18",      19, "Bank/NBFC"),
    ("Profitability (Bank): ROE 3Y ≥14%",   "roe_3y ≥ 14",      13, "Bank/NBFC"),
    ("Profitability (Bank): ROE 3Y <14%",   "roe_3y < 14",       8, "Bank/NBFC"),
    # Growth
    ("Growth: Strong 2Y compression", "pe_2y < pe_1y < pe_current",              20, "All"),
    ("Growth: 1Y compressing + 2Y flat", "pe_1y < pe_current, |pe_2y-pe_1y|<1", 14, "All"),
    ("Growth: PE flat",                  "|pe_1y - pe_current| < 1",              8, "All"),
    ("Growth: PE expanding",             "pe_1y >= pe_current + 1",              2, "All"),
]

# ─── Star Rating Bands ─────────────────────────────────────────────────────────
STAR_BANDS = [
    (5, "Strong Buy", "≥ 85 / 100"),
    (4, "Buy",        "70 – 84 / 100"),
    (3, "Hold",       "55 – 69 / 100"),
    (2, "Weak",       "40 – 54 / 100"),
    (1, "Avoid",      "< 40 / 100"),
]

BUCKET_WEIGHTS = [
    ("Trend",         20, "Price vs 50/200 DMA"),
    ("Valuation",     25, "Trailing + Forward P/E analysis"),
    ("Profitability", 35, "ROE + ROCE (3Y averages)"),
    ("Growth",        20, "Forward P/E compression"),
]


def generate_excel(out_path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Styles ─────────────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr2_fill  = PatternFill("solid", fgColor="2E6DA4")
    alt_fill   = PatternFill("solid", fgColor="EDF2F8")
    warn_fill  = PatternFill("solid", fgColor="FFF3CD")
    err_fill   = PatternFill("solid", fgColor="FFE0E0")
    ok_fill    = PatternFill("solid", fgColor="D4EDDA")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="1E3A5F")
    sub_font   = Font(bold=True, size=10, color="2E6DA4")
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap       = Alignment(wrap_text=True, vertical="top")

    def style_header_row(ws, row, cols, fill=hdr_fill):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = hdr_font
            cell.alignment = wrap
            cell.border = border

    def apply_row(ws, row_idx, values, fill=None):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.alignment = wrap
            cell.border = border
            if fill:
                cell.fill = fill

    # ── Sheet 1: Data Requirements ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Data Requirements"
    ws1.freeze_panes = "A3"

    # Title
    ws1.merge_cells("A1:N1")
    ws1["A1"] = f"Stock Rating Engine — Data Requirements  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    headers = [
        "Field Name", "DB Table", "DB Column", "Data Source",
        "Update Frequency", "Mandatory?", "Scoring Bucket", "Max Pts",
        "Sector Applicability", "Field Type", "Missing Default",
        "Missing Impact", "Ignored for Banks?", "Notes"
    ]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=2, column=c, value=h)
    style_header_row(ws1, 2, len(headers))
    ws1.row_dimensions[2].height = 24

    bucket_colors = {
        "Trend":         PatternFill("solid", fgColor="E3F0FF"),
        "Valuation":     PatternFill("solid", fgColor="FFF9E3"),
        "Valuation + Growth": PatternFill("solid", fgColor="FFF0CC"),
        "Profitability": PatternFill("solid", fgColor="E3FFE9"),
        "Growth":        PatternFill("solid", fgColor="F0E3FF"),
    }

    for i, f in enumerate(FIELDS):
        r = i + 3
        row_fill = bucket_colors.get(f["bucket"], None)
        values = [
            f["field"], f["db_table"], f["db_column"], f["source"],
            f["update_freq"], f["mandatory"], f["bucket"], f["max_pts"],
            f["sector"], f["field_type"], f["missing_default"],
            f["missing_impact"], f["bank_ignored"], f["notes"]
        ]
        apply_row(ws1, r, values, fill=row_fill)
        ws1.row_dimensions[r].height = 40

        # Colour mandatory column
        mand_cell = ws1.cell(row=r, column=6)
        if "YES" in str(f["mandatory"]):
            mand_cell.fill = ok_fill if "Optional" not in f["mandatory"] else warn_fill
        else:
            mand_cell.fill = warn_fill

        # Colour bank_ignored column
        bi_cell = ws1.cell(row=r, column=13)
        if f["bank_ignored"] == "YES":
            bi_cell.fill = warn_fill

    # Column widths
    widths = [32, 30, 20, 45, 22, 22, 22, 8, 22, 18, 42, 50, 18, 55]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Scoring Thresholds ────────────────────────────────────────────
    ws2 = wb.create_sheet("Scoring Thresholds")
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "Scoring Thresholds — All Buckets"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28

    th_headers = ["Condition Name", "Logic", "Points", "Sector"]
    for c, h in enumerate(th_headers, 1):
        ws2.cell(row=2, column=c, value=h)
    style_header_row(ws2, 2, len(th_headers))

    for i, (name, logic, pts, sector) in enumerate(THRESHOLDS):
        r = i + 3
        fill = (alt_fill if i % 2 == 0 else None)
        apply_row(ws2, r, [name, logic, pts, sector], fill=fill)

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 20

    # ── Sheet 3: Star Bands & Weights ──────────────────────────────────────────
    ws3 = wb.create_sheet("Star Bands & Weights")

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "Star Rating Bands"
    ws3["A1"].font = title_font

    star_hdr = ["Stars", "Label", "Score Range"]
    for c, h in enumerate(star_hdr, 1):
        ws3.cell(row=2, column=c, value=h)
    style_header_row(ws3, 2, 3)

    star_colors = ["006400", "228B22", "DAA520", "CC5500", "8B0000"]
    for i, (stars, label, rng) in enumerate(STAR_BANDS):
        r = i + 3
        stars_str = "★" * stars + "☆" * (5 - stars)
        ws3.cell(row=r, column=1, value=stars_str).font = Font(color=star_colors[5 - stars - 1] if (5 - stars - 1) >= 0 else star_colors[0], bold=True)
        ws3.cell(row=r, column=2, value=label)
        ws3.cell(row=r, column=3, value=rng)
        for c in range(1, 4):
            ws3.cell(row=r, column=c).border = border
            ws3.cell(row=r, column=c).alignment = wrap

    ws3.cell(row=10, column=1, value="Bucket Weights").font = sub_font
    ws3.merge_cells("A10:D10")
    bw_hdr = ["Bucket", "Max Points", "Weight %", "What it measures"]
    for c, h in enumerate(bw_hdr, 1):
        ws3.cell(row=11, column=c, value=h)
    style_header_row(ws3, 11, 4, fill=hdr2_fill)

    for i, (b, pts, desc) in enumerate(BUCKET_WEIGHTS):
        r = i + 12
        apply_row(ws3, r, [b, pts, f"{pts}%", desc])

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 40

    wb.save(out_path)
    print(f"[Excel] Saved: {out_path}")


def generate_pdf(out_path: str):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, PageBreak
        )
    except ImportError:
        print("Installing reportlab...")
        os.system("pip install reportlab -q")
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, PageBreak
        )

    NAVY   = colors.HexColor("#1E3A5F")
    BLUE   = colors.HexColor("#2E6DA4")
    LIGHT  = colors.HexColor("#EDF2F8")
    GREEN  = colors.HexColor("#D4EDDA")
    YELLOW = colors.HexColor("#FFF3CD")
    RED    = colors.HexColor("#FFE0E0")
    TREND  = colors.HexColor("#E3F0FF")
    VALUAT = colors.HexColor("#FFF9E3")
    VGROW  = colors.HexColor("#FFF0CC")
    PROFIT = colors.HexColor("#E3FFE9")
    GROWTH = colors.HexColor("#F0E3FF")

    BUCKET_BG = {
        "Trend":               TREND,
        "Valuation":           VALUAT,
        "Valuation + Growth":  VGROW,
        "Profitability":       PROFIT,
        "Growth":              GROWTH,
    }

    doc = SimpleDocTemplate(
        out_path,
        pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=14*mm, bottomMargin=14*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Normal"],
                                 fontSize=16, fontName="Helvetica-Bold",
                                 textColor=NAVY, spaceAfter=6)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontSize=10, fontName="Helvetica-Bold",
                                 textColor=BLUE, spaceAfter=4)
    cell_style  = ParagraphStyle("cell", parent=styles["Normal"],
                                 fontSize=7, leading=9)
    small_style = ParagraphStyle("small", parent=styles["Normal"],
                                 fontSize=6.5, leading=8)

    story = []

    # ─── Page 1: Main Data Requirements Table ──────────────────────────────────
    story.append(Paragraph("Stock Rating Engine — Data Requirements", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        "Based on: app/services/rating_engine.py", styles["Normal"]))
    story.append(Spacer(1, 6*mm))

    # Bucket weight summary
    wt_data = [["Bucket", "Max Pts", "Weight", "Measures"]] + \
               [[b, str(pts), f"{pts}%", desc] for b, pts, desc in BUCKET_WEIGHTS]
    wt_tbl = Table(wt_data, colWidths=[60*mm, 20*mm, 20*mm, 100*mm])
    wt_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(wt_tbl)
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=NAVY))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("Field-Level Data Requirements", sub_style))

    col_headers = [
        Paragraph("<b>Field</b>", cell_style),
        Paragraph("<b>DB Table.Column</b>", cell_style),
        Paragraph("<b>Source</b>", cell_style),
        Paragraph("<b>Update Freq</b>", cell_style),
        Paragraph("<b>Mandatory</b>", cell_style),
        Paragraph("<b>Bucket / Pts</b>", cell_style),
        Paragraph("<b>Sector</b>", cell_style),
        Paragraph("<b>Type</b>", cell_style),
        Paragraph("<b>Missing Default</b>", cell_style),
        Paragraph("<b>Missing Impact</b>", cell_style),
        Paragraph("<b>Bank Ignored?</b>", cell_style),
    ]

    col_w = [38, 30, 52, 22, 20, 22, 22, 18, 40, 50, 17]  # mm
    col_w_pts = [w * mm for w in col_w]

    tbl_data = [col_headers]
    tbl_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 7),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("ROWHEIGHT",  (0, 0), (-1, -1), 30),
    ]

    for i, f in enumerate(FIELDS):
        bg = BUCKET_BG.get(f["bucket"], colors.white)
        mand_bg = GREEN if "YES" in str(f["mandatory"]) and "Optional" not in f["mandatory"] \
                  else YELLOW
        bi_bg = YELLOW if f["bank_ignored"] == "YES" else bg

        row = [
            Paragraph(f["field"], small_style),
            Paragraph(f"{f['db_table']}<br/><i>{f['db_column']}</i>", small_style),
            Paragraph(f["source"], small_style),
            Paragraph(f["update_freq"], small_style),
            Paragraph(f["mandatory"], small_style),
            Paragraph(f"{f['bucket']}<br/><b>max {f['max_pts']} pts</b>", small_style),
            Paragraph(f["sector"], small_style),
            Paragraph(f["field_type"], small_style),
            Paragraph(f["missing_default"], small_style),
            Paragraph(f["missing_impact"], small_style),
            Paragraph(f["bank_ignored"], small_style),
        ]
        tbl_data.append(row)
        r = i + 1
        tbl_styles.append(("BACKGROUND", (0, r), (-1, r), bg))
        tbl_styles.append(("BACKGROUND", (4, r), (4, r), mand_bg))
        tbl_styles.append(("BACKGROUND", (10, r), (10, r), bi_bg))

    main_tbl = Table(tbl_data, colWidths=col_w_pts, repeatRows=1)
    main_tbl.setStyle(TableStyle(tbl_styles))
    story.append(main_tbl)

    # ─── Page 2: Scoring Thresholds ────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Scoring Thresholds — Complete Reference", title_style))
    story.append(Spacer(1, 4*mm))

    th_hdr = [["Condition", "Logic / Criterion", "Points Awarded", "Sector"]]
    th_rows = [[n, l, str(p), s] for n, l, p, s in THRESHOLDS]
    th_data = th_hdr + th_rows

    th_tbl = Table(th_data, colWidths=[100*mm, 85*mm, 30*mm, 40*mm])
    th_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]
    prev_bucket = ""
    for i, (n, l, p, s) in enumerate(THRESHOLDS):
        r = i + 1
        bucket_key = n.split(":")[0].strip()
        if "Trend" in bucket_key:     th_style_cmds.append(("BACKGROUND", (0, r), (-1, r), TREND))
        elif "Valuation" in bucket_key: th_style_cmds.append(("BACKGROUND", (0, r), (-1, r), VALUAT))
        elif "Growth" in bucket_key:  th_style_cmds.append(("BACKGROUND", (0, r), (-1, r), GROWTH))
        elif "Profitability" in bucket_key and "Bank" in n:
            th_style_cmds.append(("BACKGROUND", (0, r), (-1, r), YELLOW))
        elif "Profitability" in bucket_key:
            th_style_cmds.append(("BACKGROUND", (0, r), (-1, r), PROFIT))

    th_tbl.setStyle(TableStyle(th_style_cmds))
    story.append(th_tbl)

    # ─── Page 3: Star Bands + Missing Data Guide ────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Star Bands, Missing Data Handling & Debugging Guide", title_style))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("Star Rating Bands", sub_style))
    sb_data = [["Stars", "Label", "Score Range", "Investment Signal"]] + [
        ["★" * s + "☆" * (5 - s), lbl, rng,
         "Strong accumulate" if s == 5 else
         "Buy on dips" if s == 4 else
         "Hold / monitor" if s == 3 else
         "Consider reducing" if s == 2 else
         "Exit / avoid"] for s, lbl, rng in STAR_BANDS
    ]
    sb_tbl = Table(sb_data, colWidths=[40*mm, 40*mm, 40*mm, 80*mm])
    sb_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [GREEN, GREEN, YELLOW, RED, RED]),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
    ]))
    story.append(sb_tbl)
    story.append(Spacer(1, 6*mm))

    story.append(Paragraph("Mandatory vs Optional Fields — Quick Reference", sub_style))
    mand_data = [["Category", "Fields", "If Missing"]] + [
        ["MUST HAVE (breaks scoring)",
         "price, ma50, ma200, eps, pe",
         "Trend=4, Valuation Part B only, Growth=2"],
        ["STRONGLY RECOMMENDED",
         "forward_eps, earnings_growth, roe_3y_avg, roce_3y_avg",
         "Scores are conservative; stocks appear weaker than they are"],
        ["OPTIONAL (nice-to-have)",
         "eps_growth (user-set), roe_current (placeholder = 3Y avg)",
         "Minimal impact if Yahoo earnings_growth data is present"],
        ["BANK/NBFC IGNORED",
         "roce_3y_avg, roce_current",
         "Not penalised — points redistributed to ROE"],
        ["SECTOR FLAG",
         "is_bank (in profitability table)",
         "Default = False → bank treated as non-financial → loses up to 12 profitability pts"],
    ]
    mand_tbl = Table(mand_data, colWidths=[70*mm, 100*mm, 85*mm])
    mand_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(mand_tbl)
    story.append(Spacer(1, 6*mm))

    story.append(Paragraph("Common Reasons for Incorrect Ratings", sub_style))
    debug_data = [["Symptom", "Likely Cause", "Fix"]] + [
        ["All stocks rated 1–2★",
         "profitability data not populated (roe_3y_avg = NULL)",
         "Run compute_profitability_summary.py for all symbols"],
        ["Banks rated very low",
         "is_bank flag not set → ROCE=0 factored in",
         "Run profitability script with sector detection enabled"],
        ["Trend always 4/20",
         "ma50 or ma200 missing in price_cache",
         "Ensure yfinance historical data fetch is working"],
        ["Valuation Part A always 2",
         "EPS = 0 or NULL → pe_1y/pe_2y cannot be derived",
         "Ensure EPS is populated and > 0 for all scrips"],
        ["Growth always 2",
         "pe_current = NULL or EPS = 0",
         "Same fix as Valuation; EPS is the root dependency"],
        ["NBFC treated as non-financial",
         "Symbol not in hard-coded NBFC list in rating_engine.py",
         "Add symbol to NBFC list in compute_and_store_rating()"],
    ]
    dbg_tbl = Table(debug_data, colWidths=[70*mm, 95*mm, 90*mm])
    dbg_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(dbg_tbl)

    doc.build(story)
    print(f"[PDF]   Saved: {out_path}")


if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    pdf_path   = os.path.join(base, "rating_data_requirements.pdf")
    excel_path = os.path.join(base, "rating_data_requirements.xlsx")

    generate_excel(excel_path)
    generate_pdf(pdf_path)

    print("\nDone. Files written to backend/:")
    print(f"  PDF:   {pdf_path}")
    print(f"  Excel: {excel_path}")
