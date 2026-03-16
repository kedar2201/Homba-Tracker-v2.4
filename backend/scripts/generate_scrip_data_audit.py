"""
Per-Scrip Data Audit Report Generator
Queries the live database and produces:
  - scrip_data_audit.xlsx  (per-scrip field presence heatmap)
  - scrip_data_audit.pdf   (printable summary)

Run from backend/ directory:  python generate_scrip_data_audit.py
"""

import os
import sys

# ── DB bootstrap ────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from app.database import SessionLocal
from app.models.market_data import PriceCache
from app.models.profitability_metrics import StockProfitabilitySummary
from app.models.rating import StockRatingSummary

from datetime import datetime

# ── Helpers ─────────────────────────────────────────────────────────────────
def _ok(val):
    """True if value is non-None and non-zero (for numeric fields)."""
    if val is None:
        return False
    if isinstance(val, (int, float)) and val == 0:
        return False
    return True

def _fmt(val, decimals=2):
    if val is None:
        return "—"
    if isinstance(val, float):
        return f"{val:.{decimals}f}"
    return str(val)

NBFC_LIST = {"BAJAJFIN", "BAJFINANCE", "CFL", "HDFC", "LIC",
             "LICHSGFIN", "MUTHOOTFIN", "PFC", "REC"}

MANDATORY_FIELDS = ["yahoo_symbol", "price", "ma50", "ma200", "eps", "pe"]
IMPORTANT_FIELDS = ["forward_eps", "earnings_growth"]
PROFITABILITY_FIELDS = ["roe_3y_avg", "roce_3y_avg"]

# ── Pull data ────────────────────────────────────────────────────────────────
def fetch_data():
    db = SessionLocal()
    try:
        price_rows = {r.symbol: r for r in db.query(PriceCache).all()}
        prof_rows  = {r.scrip_code: r for r in db.query(StockProfitabilitySummary).all()}
        rating_rows = {r.scrip_code: r for r in db.query(StockRatingSummary).all()}

        all_codes = sorted(set(list(price_rows.keys()) + list(prof_rows.keys())))

        records = []
        for code in all_codes:
            pc  = price_rows.get(code)
            pr  = prof_rows.get(code)
            rt  = rating_rows.get(code)

            is_bank_flag = False
            # Priority 1: Database flag
            if pr is not None:
                is_bank_flag = pr.is_bank
            # Priority 2: Name-based fallback (if DB entry doesn't exist yet)
            elif code in NBFC_LIST or any(kw in code.upper() for kw in ["BANK", "FINANCE"]):
                is_bank_flag = True

            sector = "Bank/NBFC" if is_bank_flag else "Non-Financial"

            rec = {
                "scrip_code":     code,
                "sector":         sector,
                # PriceCache fields
                "price":          pc.price      if pc else None,
                "ma50":           pc.ma50       if pc else None,
                "ma200":          pc.ma200      if pc else None,
                "eps":            pc.eps        if pc else None,
                "pe":             pc.pe         if pc else None,
                "forward_eps":    pc.forward_eps if pc else None,
                "earnings_growth":pc.earnings_growth if pc else None,
                "eps_growth":     pc.eps_growth if pc else None,
                "yahoo_symbol":   pc.yahoo_symbol if pc else None,
                "yahoo_symbol_locked": pc.yahoo_symbol_locked if pc else False,
                # Profitability fields
                "roe_3y_avg":     pr.roe_3y_avg  if pr else None,
                "roce_3y_avg":    pr.roce_3y_avg if pr else None,
                "is_bank":        pr.is_bank     if pr else None,
                # Rating output
                "final_score":    rt.final_score    if rt else None,
                "star_rating":    rt.star_rating    if rt else None,
                "trend_score":    rt.trend_score    if rt else None,
                "valuation_score":rt.valuation_score if rt else None,
                "profitability_score": rt.profitability_score if rt else None,
                "growth_score":   rt.growth_score   if rt else None,
                "last_updated":   rt.last_updated.strftime("%Y-%m-%d") if (rt and rt.last_updated) else None,
            }

            # Data completeness score (out of 10 key fields)
            filled = sum([
                _ok(rec["yahoo_symbol"]),
                _ok(rec["price"]), _ok(rec["ma50"]), _ok(rec["ma200"]),
                _ok(rec["eps"]), _ok(rec["pe"]),
                _ok(rec["forward_eps"]), _ok(rec["earnings_growth"]),
                _ok(rec["roe_3y_avg"]),
                False if (sector == "Bank/NBFC") else _ok(rec["roce_3y_avg"]),
            ])
            total_fields = 10 if sector != "Bank/NBFC" else 9
            rec["data_completeness"] = f"{filled}/{total_fields}"
            rec["completeness_pct"]  = round(filled / total_fields * 100)

            # Identify what's missing
            missing = []
            for f in MANDATORY_FIELDS:
                if not _ok(rec[f]):
                    missing.append(f"❌ {f}")
            for f in IMPORTANT_FIELDS:
                if not _ok(rec[f]):
                    missing.append(f"⚠ {f}")
            if not _ok(rec["roe_3y_avg"]):
                missing.append("⚠ roe_3y_avg")
            if sector != "Bank/NBFC" and not _ok(rec["roce_3y_avg"]):
                missing.append("⚠ roce_3y_avg")
            rec["missing_fields"] = ", ".join(missing) if missing else "✓ Complete"

            records.append(rec)

        return records
    finally:
        db.close()


# ── Excel ─────────────────────────────────────────────────────────────────────
def generate_excel(records, out_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Scrip Data Audit"
    ws.freeze_panes = "C3"

    # Styles
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    ok_fill    = PatternFill("solid", fgColor="C6EFCE")
    warn_fill  = PatternFill("solid", fgColor="FFEB9C")
    err_fill   = PatternFill("solid", fgColor="FFC7CE")
    na_fill    = PatternFill("solid", fgColor="D9D9D9")
    bank_fill  = PatternFill("solid", fgColor="DAE8FC")
    hdr_font   = Font(bold=True, color="FFFFFF", size=9)
    title_font = Font(bold=True, size=14, color="1E3A5F")
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap       = Alignment(wrap_text=True, vertical="top", horizontal="center")
    left       = Alignment(wrap_text=False, vertical="top", horizontal="left")

    # Title row
    ws.merge_cells("A1:W1")
    ws["A1"] = f"Per-Scrip Data Audit  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    headers = [
        "Scrip Code", "Sector",
        # PriceCache
        "Price", "MA50", "MA200", "EPS", "P/E",
        "Fwd EPS (Yahoo)", "Earn Growth (Yahoo)", "EPS Growth (User)",
        # Profitability
        "ROE 3Y Avg", "ROCE 3Y Avg", "is_bank",
        # Rating output
        "Final Score", "Stars", "Trend", "Valuation", "Profitability", "Growth",
        "Last Computed",
        # Summary
        "Yahoo Mapping", "Locked?",
        "Completeness", "%", "Missing / Warnings"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = wrap
        cell.border = border
    ws.row_dimensions[2].height = 30

    # Data rows
    for i, rec in enumerate(records):
        r = i + 3
        is_bank = rec["sector"] == "Bank/NBFC"

        def cell_val_fill(field, value, fmt_val):
            if is_bank and field in ["roce_3y_avg"]:
                return fmt_val, na_fill  # N/A for banks
            if value is None or (isinstance(value, (int, float)) and value == 0 and field not in ["is_bank"]):
                return fmt_val, err_fill if field in MANDATORY_FIELDS + PROFITABILITY_FIELDS else warn_fill
            return fmt_val, ok_fill

        row_data = [
            (rec["scrip_code"],  None,  None),
            (rec["sector"],      None,  bank_fill if is_bank else None),
        ]

        numeric_fields = [
            ("price",           rec["price"],           _fmt(rec["price"])),
            ("ma50",            rec["ma50"],            _fmt(rec["ma50"])),
            ("ma200",           rec["ma200"],           _fmt(rec["ma200"])),
            ("eps",             rec["eps"],             _fmt(rec["eps"])),
            ("pe",              rec["pe"],              _fmt(rec["pe"])),
            ("forward_eps",     rec["forward_eps"],     _fmt(rec["forward_eps"])),
            ("earnings_growth", rec["earnings_growth"], _fmt(rec["earnings_growth"], 4)),
            ("eps_growth",      rec["eps_growth"],      _fmt(rec["eps_growth"])),
            ("roe_3y_avg",      rec["roe_3y_avg"],      _fmt(rec["roe_3y_avg"])),
            ("roce_3y_avg",     rec["roce_3y_avg"],     _fmt(rec["roce_3y_avg"])),
            ("is_bank",         rec["is_bank"],         "Yes" if rec["is_bank"] else ("No" if rec["is_bank"] is not None else "—")),
        ]

        rating_fields = [
            rec["final_score"], "★" * (rec["star_rating"] or 0) if rec["star_rating"] else "—",
            rec["trend_score"], rec["valuation_score"],
            rec["profitability_score"], rec["growth_score"],
            rec["last_updated"] or "—",
            rec["yahoo_symbol"] or "—",
            "Yes" if rec["yahoo_symbol_locked"] else "No",
        ]

        summary_fields = [
            rec["data_completeness"],
            f"{rec['completeness_pct']}%",
            rec["missing_fields"],
        ]

        # Write scrip + sector
        for c, (val, fmt, fill) in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = left
            cell.border = border
            if fill:
                cell.fill = fill

        # Write numeric fields with conditional colour
        col = 3
        for field, raw, fmt_v in numeric_fields:
            display, fill = cell_val_fill(field, raw, fmt_v)
            cell = ws.cell(row=r, column=col, value=display)
            cell.fill = fill
            cell.alignment = wrap
            cell.border = border
            col += 1

        # Write rating fields
        for val in rating_fields:
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = wrap
            cell.border = border
            if rec["final_score"] is not None and col == 14:  # score col
                score = rec["final_score"]
                cell.fill = (ok_fill if score >= 70 else warn_fill if score >= 50 else err_fill)
            col += 1

        # Write summary fields
        for idx, val in enumerate(summary_fields):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            if idx == 1:  # % column
                pct = rec["completeness_pct"]
                cell.fill = ok_fill if pct >= 90 else warn_fill if pct >= 60 else err_fill
                cell.alignment = wrap
            elif idx == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if "❌" in str(val):
                    cell.fill = err_fill
                elif "✓" in str(val):
                    cell.fill = ok_fill
                else:
                    cell.fill = warn_fill
            else:
                cell.alignment = wrap
            col += 1

    # Column widths
    widths = [16, 13, 10, 10, 10, 10, 10, 14, 16, 14, 12, 12, 9, 12, 8, 10, 12, 14, 10, 14, 15, 10, 12, 8, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Add legend sheet
    ws2 = wb.create_sheet("Legend")
    ws2["A1"] = "Color Legend"
    ws2["A1"].font = Font(bold=True, size=12, color="1E3A5F")
    legend = [
        ("Green",  ok_fill,   "Field is populated with a valid non-zero value"),
        ("Yellow", warn_fill, "Field is NULL/0 but optional — engine has fallbacks"),
        ("Red",    err_fill,  "Field is NULL/0 and MANDATORY — scoring will be degraded"),
        ("Grey",   na_fill,   "Field is N/A for this sector (e.g. ROCE skipped for Banks)"),
        ("Blue",   bank_fill, "Bank or NBFC sector"),
    ]
    for r, (label, fill, desc) in enumerate(legend, 3):
        c1 = ws2.cell(row=r, column=1, value=label)
        c1.fill = fill
        c1.border = border
        c2 = ws2.cell(row=r, column=2, value=desc)
        c2.alignment = Alignment(wrap_text=True)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 65

    wb.save(out_path)
    print(f"[Excel] Saved: {out_path}  ({len(records)} scrips)")


# ── PDF ───────────────────────────────────────────────────────────────────────
def generate_pdf(records, out_path):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, HRFlowable)
    except ImportError:
        os.system("pip install reportlab -q")
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, HRFlowable)

    NAVY   = colors.HexColor("#1E3A5F")
    GREEN  = colors.HexColor("#C6EFCE")
    YELLOW = colors.HexColor("#FFEB9C")
    RED    = colors.HexColor("#FFC7CE")
    GREY   = colors.HexColor("#D9D9D9")
    BANK   = colors.HexColor("#DAE8FC")

    doc = SimpleDocTemplate(out_path, pagesize=landscape(A4),
                             leftMargin=10*mm, rightMargin=10*mm,
                             topMargin=12*mm, bottomMargin=12*mm)

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Normal"], fontSize=14,
                              fontName="Helvetica-Bold", textColor=NAVY, spaceAfter=4)
    small_s = ParagraphStyle("s", parent=styles["Normal"], fontSize=6.5, leading=8)
    tiny_s  = ParagraphStyle("tiny", parent=styles["Normal"], fontSize=6, leading=7.5)

    story = []
    story.append(Paragraph("Per-Scrip Data Audit — Stock Rating Engine", title_s))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Total scrips: {len(records)}", styles["Normal"]))
    story.append(Spacer(1, 4*mm))

    # Summary stats
    total = len(records)
    full  = sum(1 for r in records if r["completeness_pct"] >= 90)
    partial = sum(1 for r in records if 50 <= r["completeness_pct"] < 90)
    poor  = sum(1 for r in records if r["completeness_pct"] < 50)
    rated = sum(1 for r in records if r["star_rating"] is not None)

    stat_data = [["Total Scrips", "Fully Complete (≥90%)", "Partial (50–89%)", "Poor (<50%)", "Have Ratings"]] + \
                [[str(total), str(full), str(partial), str(poor), str(rated)]]
    stat_tbl = Table(stat_data, colWidths=[40*mm]*5)
    stat_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (1, 1), (1, 1), GREEN),
        ("BACKGROUND", (2, 1), (2, 1), YELLOW),
        ("BACKGROUND", (3, 1), (3, 1), RED),
        ("FONTNAME",   (0, 1), (-1, 1), "Helvetica-Bold"),
    ]))
    story.append(stat_tbl)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=NAVY))
    story.append(Spacer(1, 3*mm))

    # Main table
    col_headers = [
        "Scrip", "Sector", "Price", "MA50", "MA200", "EPS", "P/E",
        "Fwd EPS", "Earn\nGrowth", "EPS\nGrowth",
        "ROE\n3Y", "ROCE\n3Y", "Score", "Stars",
        "T", "V", "P", "G",
        "Computed", "Yahoo Map", "Lock", "Complete", "Missing / Warnings"
    ]
    col_w = [18, 13, 10, 10, 10, 9, 9, 10, 10, 9, 9, 9, 11, 10, 7, 7, 7, 7, 14, 14, 8, 12, 50]
    col_w_pts = [w * mm for w in col_w]

    def cell(v, style=small_s):
        return Paragraph(str(v) if v is not None else "—", style)

    tbl_data = [[Paragraph(f"<b>{h}</b>", small_s) for h in col_headers]]
    tbl_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 6.5),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]

    def field_color(field, val, is_bank_row):
        if is_bank_row and field == "roce_3y_avg":
            return GREY
        if val is None or (isinstance(val, (int, float)) and val == 0):
            if field in MANDATORY_FIELDS + PROFITABILITY_FIELDS:
                return RED
            return YELLOW
        return GREEN

    for i, rec in enumerate(records):
        row_num = i + 1
        is_bank = rec["sector"] == "Bank/NBFC"

        stars_str = "★" * (rec["star_rating"] or 0) if rec["star_rating"] else "—"

        numeric_vals = [
            ("price",           rec["price"],            _fmt(rec["price"])),
            ("ma50",            rec["ma50"],             _fmt(rec["ma50"])),
            ("ma200",           rec["ma200"],            _fmt(rec["ma200"])),
            ("eps",             rec["eps"],              _fmt(rec["eps"])),
            ("pe",              rec["pe"],               _fmt(rec["pe"])),
            ("forward_eps",     rec["forward_eps"],      _fmt(rec["forward_eps"])),
            ("earnings_growth", rec["earnings_growth"],  _fmt(rec["earnings_growth"], 4)),
            ("eps_growth",      rec["eps_growth"],       _fmt(rec["eps_growth"])),
            ("roe_3y_avg",      rec["roe_3y_avg"],       _fmt(rec["roe_3y_avg"])),
            ("roce_3y_avg",     rec["roce_3y_avg"],      _fmt(rec["roce_3y_avg"])),
        ]

        row = [
            cell(rec["scrip_code"]),
            cell(rec["sector"]),
        ]
        for field, raw, fmt_v in numeric_vals:
            row.append(cell(fmt_v))

        row += [
            cell(_fmt(rec["final_score"], 1) if rec["final_score"] else "—"),
            cell(stars_str),
            cell(rec["trend_score"] or "—"),
            cell(rec["valuation_score"] or "—"),
            cell(rec["profitability_score"] or "—"),
            cell(rec["growth_score"] or "—"),
            cell(rec["last_updated"] or "—"),
            cell(rec["yahoo_symbol"] or "—", tiny_s),
            cell("Yes" if rec["yahoo_symbol_locked"] else "No"),
            cell(rec["data_completeness"]),
            Paragraph(rec["missing_fields"], tiny_s),
        ]

        tbl_data.append(row)

        # Row background for sector
        row_fill = BANK if is_bank else (colors.white if i % 2 == 0 else colors.HexColor("#F8FAFF"))
        tbl_styles.append(("BACKGROUND", (0, row_num), (1, row_num), row_fill))

        # Per-cell coloring for numeric fields (cols 2–11)
        for j, (field, raw, _) in enumerate(numeric_vals):
            c_idx = 2 + j
            bg = field_color(field, raw, is_bank)
            tbl_styles.append(("BACKGROUND", (c_idx, row_num), (c_idx, row_num), bg))

        # Score column
        if rec["final_score"] is not None:
            sc = rec["final_score"]
            sg = GREEN if sc >= 70 else YELLOW if sc >= 50 else RED
            tbl_styles.append(("BACKGROUND", (12, row_num), (12, row_num), sg))

        # Missing col (col 22)
        mf = rec["missing_fields"]
        mg = GREEN if "✓" in mf else RED if "❌" in mf else YELLOW
        tbl_styles.append(("BACKGROUND", (22, row_num), (22, row_num), mg))

    main_tbl = Table(tbl_data, colWidths=col_w_pts, repeatRows=1)
    main_tbl.setStyle(TableStyle(tbl_styles))
    story.append(main_tbl)

    doc.build(story)
    print(f"[PDF]   Saved: {out_path}  ({len(records)} scrips)")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Fetching scrip data from database...")
    records = fetch_data()
    print(f"Found {len(records)} scrips.")

    base  = os.path.dirname(os.path.abspath(__file__))
    xlsx  = os.path.join(base, "scrip_data_audit.xlsx")
    pdf   = os.path.join(base, "scrip_data_audit.pdf")

    generate_excel(records, xlsx)
    generate_pdf(records, pdf)

    print(f"\nDone.")
    print(f"  Excel: {xlsx}")
    print(f"  PDF:   {pdf}")
